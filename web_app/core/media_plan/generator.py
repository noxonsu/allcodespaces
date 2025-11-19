"""
Services responsible for turning campaigns into Excel media plans.

CHANGE: Added reusable generator with deterministic template.
WHY: Issues #49-50 require Excel export, history logging, and error handling.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Iterable, Sequence

from django.db.models import QuerySet, Sum, Count
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.media_plan.template import MediaPlanTemplate, MEDIA_PLAN_TEMPLATE
from core.models import Campaign, CampaignChannel


class MediaPlanGenerationError(Exception):
    """Generic exception when we cannot build a media plan."""


@dataclass(frozen=True)
class MediaPlanGenerationResult:
    """Result of generator run."""

    filename: str
    content: bytes
    rows: Sequence[dict]
    totals: dict


class MediaPlanGenerator:
    """Excel builder for campaign media plans."""

    def __init__(self, template: MediaPlanTemplate | None = None):
        self.template = template or MEDIA_PLAN_TEMPLATE

    def generate(self, campaigns: QuerySet[Campaign] | Iterable[Campaign]) -> MediaPlanGenerationResult:
        """Return Excel bytes and meta information for provided campaigns."""
        campaign_list = self._prepare_campaigns(campaigns)
        if not campaign_list:
            raise MediaPlanGenerationError("Нет кампаний для генерации медиаплана")

        rows = self._build_rows(campaign_list)
        totals = self._calculate_totals(campaign_list, rows)
        workbook_bytes = self._render_workbook(rows)
        filename = f"media_plan_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return MediaPlanGenerationResult(
            filename=filename,
            content=workbook_bytes,
            rows=rows,
            totals=totals,
        )

    def _prepare_campaigns(self, campaigns: QuerySet[Campaign] | Iterable[Campaign]) -> list[Campaign]:
        if isinstance(campaigns, QuerySet):
            qs = campaigns.select_related("message").order_by("start_date", "name")
            return list(qs)
        return list(campaigns)

    def _build_rows(self, campaigns: Sequence[Campaign]) -> list[dict]:
        """Aggregate per-campaign values for template columns."""
        stats = (
            CampaignChannel.objects.filter(campaign__in=campaigns)
            .values("campaign_id")
            .annotate(
                channels_count=Count("id", distinct=True),
                impressions_plan=Coalesce(Sum("impressions_plan"), 0),
                impressions_fact=Coalesce(Sum("impressions_fact"), 0),
                clicks=Coalesce(Sum("clicks"), 0),
            )
        )
        stats_map = {row["campaign_id"]: row for row in stats}

        rows: list[dict] = []
        for campaign in campaigns:
            stat = stats_map.get(campaign.id, {})
            message = campaign.message
            rows.append(
                {
                    "campaign_id": str(campaign.id),
                    "campaign_name": campaign.name,
                    "campaign_status": campaign.get_status_display(),
                    "campaign_format": campaign.get_format_display(),
                    "client": campaign.client,
                    "brand": campaign.brand,
                    "budget": campaign.budget,
                    "start_date": campaign.start_date,
                    "finish_date": campaign.finish_date,
                    "slot_publication_at": campaign.slot_publication_at,
                    "channels_count": stat.get("channels_count", 0),
                    "impressions_plan": stat.get("impressions_plan", 0),
                    "impressions_fact": stat.get("impressions_fact", 0),
                    "clicks": stat.get("clicks", 0),
                    "creative_title": getattr(message, "title", "") or getattr(message, "name", ""),
                }
            )
        return rows

    def _calculate_totals(self, campaigns: Sequence[Campaign], rows: Sequence[dict]) -> dict:
        total_budget = sum((campaign.budget for campaign in campaigns), Decimal("0"))
        total_impressions_plan = sum((row["impressions_plan"] for row in rows), 0)
        total_impressions_fact = sum((row["impressions_fact"] for row in rows), 0)
        total_clicks = sum((row["clicks"] for row in rows), 0)
        total_channels = sum((row["channels_count"] for row in rows), 0)
        return {
            "campaigns": len(campaigns),
            "budget": str(total_budget),
            "impressions_plan": total_impressions_plan,
            "impressions_fact": total_impressions_fact,
            "clicks": total_clicks,
            "channels": total_channels,
        }

    def _render_workbook(self, rows: Sequence[dict]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.template.sheet_title
        worksheet.freeze_panes = "A2"

        header_font = Font(bold=True, color=self.template.header_font_color)
        header_fill = PatternFill("solid", fgColor=self.template.header_fill)
        header_alignment = Alignment(horizontal=self.template.header_alignment, vertical="center")
        body_alignment = Alignment(vertical="top", wrap_text=True)

        # Header
        for column_index, column in enumerate(self.template.columns, start=1):
            excel_column = get_column_letter(column_index)
            worksheet.column_dimensions[excel_column].width = column.width
            cell = worksheet.cell(row=1, column=column_index, value=column.title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data
        for row_index, row in enumerate(rows, start=2):
            for column_index, column in enumerate(self.template.columns, start=1):
                value = row.get(column.key)
                if isinstance(value, datetime) and timezone.is_aware(value):
                    value = timezone.localtime(value).replace(tzinfo=None)
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = body_alignment
                if column.number_format:
                    cell.number_format = column.number_format

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()
